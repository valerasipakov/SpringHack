import json
import datetime
import openpyxl
from openpyxl.styles import Alignment


class ProductInfo:

    def __init__(self, product, url, photo, feature, typeProduct):
        self.product = product
        self.url = url
        self.photo = photo
        self.feature = feature
        self.typeProduct = typeProduct


class Converter:

    def __init__(self, json_file, xlsx_file):
        self.json_file = json_file
        self.xlsx_file = xlsx_file

    def run(self):
        with open(self.json_file, 'r', encoding='utf-8') as inputFile:
            parse_text = json.load(inputFile)

        objtmstmp = float(parse_text['дата сбора'])
        self.xlsx_file += datetime.datetime.fromtimestamp(objtmstmp).strftime("%Y-%m-%d") + ".xlsx"

        for txt in parse_text['content']:
            instance = ProductInfo(txt['название'], txt['ссылка'], txt['изображение'], txt['характеристики'], txt['категория'])
            self.write(instance)

    def write(self, instance):
        try:
            excel_file = openpyxl.load_workbook(self.xlsx_file)
        except:
            excel_file = openpyxl.Workbook()
            for sheet_name in excel_file.sheetnames:
                sheet = excel_file[sheet_name]
                excel_file.remove(sheet)

        try:
            sheet = excel_file[instance.typeProduct]
        except:
            sheet = excel_file.create_sheet(instance.typeProduct)
            sheet.cell(row=1, column=1).value = "Наименование"
            sheet.cell(row=1, column=2).value = "Ссылка"
            sheet.cell(row=1, column=3).value = "Фото"
            list_feature = list(instance.feature.keys())
            for i in range(len(list_feature)):
                sheet.cell(row=1, column=i + 4).value = list_feature[i]

        line_write = sheet.max_row + 1
        sheet.cell(row=line_write, column=1).value = instance.product
        sheet.cell(row=line_write, column=2).value = instance.url
        alignment = Alignment(wrap_text=True)
        sheet.cell(row=line_write, column=3).alignment = alignment
        sheet.cell(row=line_write, column=3).value = '\n'.join(instance.photo)
        list_feature = list(instance.feature.keys())
        for i in range(len(list_feature)):
            tmp = instance.feature[list_feature[i]]
            if type(tmp) != list:
                sheet.cell(row=line_write, column=4 + i).value = tmp
            else:
                sheet.cell(row=line_write, column=4 + i).value = ' '.join(tmp)
        excel_file.save(self.xlsx_file)

if __name__ == "__main__":
    convert = Converter("data/catalog_v9.json", "Hb")
    convert.run()